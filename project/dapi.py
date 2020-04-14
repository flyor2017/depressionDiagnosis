from openpyxl import load_workbook
import json
import os

from .graph import RuleNode
from .graph import SymptomNode
from .graph import DiseaseNode
from .graph import ComputableNode
from .treatment import Treatment

__author__ = "flyor"

r"""
该类主要的功能是封装整个模型，对外提供访问接口。
"""


class Graph(object):

    diseaseFile = "resources/知识图谱-核心实体关系v1.3.xlsx"
    symptomFile = "resources/知识图谱-核心实体关系v1.3.xlsx"
    ruleFile = "resources/rule.json"

    def __init__(self):
        super().__init__()

        # 整个图的核心节点
        self._root = DiseaseNode()
        self._root.name = "rootNode"

        # 所有疾病节点
        self._diseaseMap = {}

        # 所有症状节点
        self._symptomMap = {}

        # 治疗方案的给出对象
        self._treatment = Treatment()

        # 拿到当前脚本文件所在的目录
        dirname = os.path.dirname(__file__)
        diseaseFile = os.path.join(dirname, Graph.diseaseFile)
        symptomFile = os.path.join(dirname, Graph.symptomFile)
        ruleFile = os.path.join(dirname, Graph.ruleFile)

        # 从资源文件夹中读取所有数据
        self.readDisease(diseaseFile, "疾病-疾病关系")
        self.readSymptom(symptomFile, "大症状-小症状关系")
        self.readSymptom(symptomFile, "小症状-小症状关系")
        self.readRule(ruleFile)

    def readDisease(self, file, sheetName):
        r"""
        从excel表中读取疾病，数据组织方式应该是三元组形式。
        格式为：疾病 sub_disease 子疾病
        这个函数并没有用到三元组的谓语，中间这一项并不会产生任何影响
        """
        wb = load_workbook(filename=file)
        sheet_ranges = wb[sheetName]

        rowNumber = 0
        for row in sheet_ranges.values:

            # 行循环从第二行开始，跳过表头字段
            rowNumber += 1
            if rowNumber == 1:
                continue

            # 跳过项数小于3的行，三元组至少要三项
            if len(row) < 3:
                continue

            # 前三项有一项为空则跳过
            for i in range(0, 3):
                if row[i] == None or len(row[i]) == 0:
                    break
            else:
                # 提取主语和宾语，即对应疾病和子疾病
                sub_str = row[0]
                obj_str = row[2]

                # 先尝试从所有疾病节点中获取该疾病节点，
                # 如果获取失败则创建该疾病节点
                sub_node = self._diseaseMap.get(sub_str, None)
                if sub_node == None:
                    sub_node = DiseaseNode(self._root)
                    sub_node.name = sub_str
                    self._diseaseMap[sub_str] = sub_node

                # 对于子疾病节点，一般来说不会存在已有节点，
                # 所以直接创建新的疾病节点
                obj_node = DiseaseNode(sub_node)
                self._diseaseMap[obj_str] = obj_node
                obj_node.name = obj_str

    def readSymptom(self, file, sheetName):
        r"""
        从excel表中读取症状，数据组织方式应该是三元组形式。
        格式为：症状 sub_symptom 子症状
        这个函数并没有用到三元组的谓语，中间这一项并不会产生任何影响
        """
        wb = load_workbook(filename=file)
        sheet_ranges = wb[sheetName]

        rowNumber = 0
        for row in sheet_ranges.values:

            # 行循环从第二行开始，跳过表头字段
            rowNumber += 1
            if rowNumber == 1:
                continue

            # 跳过项数小于3的行，三元组至少要三项
            if len(row) < 3:
                continue

            # 前三项有一项为空则跳过
            for i in range(0, 3):
                if row[i] == None or len(row[i]) == 0:
                    break
            else:
                # 提取主语和宾语，即对应症状和子症状
                sub_str = row[0]
                obj_str = row[2]

                # 先尝试从所有症状节点中获取该症状节点，
                # 如果获取失败则创建该症状节点
                sub_node = self._symptomMap.get(sub_str, None)
                if sub_node == None:
                    sub_node = SymptomNode()
                    sub_node.name = sub_str
                    self._symptomMap[sub_str] = sub_node

                # 症状对于宾语节点的做法和疾病宾语节点有所不同
                # 因为疾病每一个宾语节点（子疾病），有一个父节点（疾病）
                # 而症状的每一个宾语节点（子症状），可能有多个父节点（症状）
                # 包括症状本身，也可能从属多个诊断规则，即有多个父节点
                obj_node = self._symptomMap.get(obj_str, None)
                if obj_node == None:
                    obj_node = SymptomNode()
                    obj_node.name = obj_str
                    self._symptomMap[obj_str] = obj_node
                sub_node.addSubSymptom(obj_node)

    def readRule(self, filename):
        r"""
        该方法的作用是将诊断规则读入到图中，在没有读入诊断规则之前，
        疾病节点和症状节点是分开的，规则节点是将二者连接起来的中间节点。
        """

        # 从json文件中读取诊断规则
        jsonObj = None
        with open(filename, encoding="utf-8") as f:
            jsonObj = json.load(f)

        jsonObj = jsonObj["contents"]

        # 为每一个疾病添加诊断规则
        for term in jsonObj:

            # 对每一个疾病进行循环
            diseaseName = term["name"]
            diseaseNode = self._diseaseMap[diseaseName]
            for rule in term["rules"]:

                # 一个疾病可能有好几个规则
                # 要同时满足这些规则才能满足患有该疾病的要求
                rn = RuleNode(diseaseNode)
                symptomName = rule["symptom"]

                # 设置该规则控制的症状以及需要满足的症状的条数
                # 先获得该症状分类下的所有子症状
                subSymptoms = self._symptomMap[symptomName].getChildren()
                rn.setSymptoms(subSymptoms)
                rn.count = rule["count"]

    def printGraph(self):
        r"""
        以树形结构打印该图
        """
        self._root.printGraph()

    def setProbability(self, symptom, probability, degree=SymptomNode.NORMAL_DEGREE, negator=False):
        r"""
        设置某一症状的概率值、程度值、是否包含否定词，
        如果未找到该症状则返回False，否则返回True。
        """
        s = self._symptomMap.get(symptom, None)
        if s == None:
            return False

        # 概率值
        s.probability = probability

        # 该症状的程度
        s.degree = degree

        # 描述该症状是否用了否定词
        s.negator = negator

        return True

    def getSyntheticalProbability(self, s):
        r"""
        返回患某一疾病或症状的综合概率，
        如果没有找到该疾病或症状，则返回0
        """
        disease = self._diseaseMap.get(s, None)
        if disease != None:
            return disease.getSyntheticalProbability()
        else:
            symptom = self._symptomMap.get(s, None)
            if symptom == None:
                return 0
            else:
                return symptom.getSyntheticalProbability()

    def reset(self):
        r"""
        重置整个图结构
        """
        for i in self._symptomMap.values():
            i.degree = SymptomNode.NORMAL_DEGREE
            i.negator = False
            i.probability = ComputableNode.K

    def printTreatmentScheme(self, diseases):
        r"""
        打印建议的治疗方案
        """

        # 列表为空直接返回
        if diseases == None or len(diseases) == 0:
            return

        # 获取整个列表疾病的概率值
        arr = []
        print("-------start---------")
        for i in diseases:
            arr.append(self.getSyntheticalProbability(i))
            print("{0:　<16}{1:.4f}".format(i, arr[-1]))
        print("--------end----------")

        # 找出有可能患的疾病
        index = 0
        m = 0
        for i, v in enumerate(arr):
            if v > m:
                m = v
                index = i

        # 打印该疾病的治疗方案
        self._treatment.printScheme(diseases[index], self)
